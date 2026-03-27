import re
import os
import log
import sqlite3
from pathlib import Path
from typing import Any, Dict, List, Iterator
from openpyxl.cell import Cell
from collections import defaultdict

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook, Workbook

import hashlib
from uuid import uuid4

log.enable()
log.throttle()


SCHEMA_SQL = """
PRAGMA journal_mode = WAL;
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS materials (
    id                   TEXT PRIMARY KEY NOT NULL,

    brand                TEXT NOT NULL,
    category             TEXT NOT NULL,
    quality              TEXT,

    height_m             REAL NOT NULL,
    width_m              REAL NOT NULL,
    length_m             REAL NOT NULL,
    
    file_name            TEXT NOT NULL,
    sheet_name           TEXT NOT NULL,
    sheet_name_orderable TEXT NOT NULL,
    row                  INTEGER NOT NULL
);

CREATE TABLE IF NOT EXISTS end_of_month_material_stats (
    id                   BLOB PRIMARY KEY NOT NULL,

    material_id          TEXT NOT NULL,
    year                 INTEGER NOT NULL,
    month                INTEGER NOT NULL,

    price_purchase       REAL,
    price_m3             REAL,
    price_pcs            REAL,
    price_m2             REAL,

    qty_pcs              REAL,
    qty_m3               REAL,
    qty_m2               REAL,
    qty_eur              REAL,
    
    file_name            TEXT NOT NULL,
    sheet_name           TEXT NOT NULL,
    sheet_name_orderable TEXT NOT NULL,
    row                  INTEGER NOT NULL,

    FOREIGN KEY (material_id) REFERENCES materials(id)
);

CREATE TABLE IF NOT EXISTS sales (
    id                   BLOB PRIMARY KEY NOT NULL,

    material_id          TEXT NOT NULL,
    year                 INTEGER NOT NULL,
    month                INTEGER NOT NULL,

    sale_pcs             REAL NOT NULL,
    sale_m3              REAL NOT NULL,
    sale_eur             REAL NOT NULL,
    
    file_name            TEXT NOT NULL,
    sheet_name           TEXT NOT NULL,
    sheet_name_orderable TEXT NOT NULL,
    row                  INTEGER NOT NULL,

    FOREIGN KEY (material_id) REFERENCES materials(id)
);

CREATE TABLE IF NOT EXISTS deliveries (
    id                   BLOB PRIMARY KEY NOT NULL,

    material_id          TEXT NOT NULL,
    year                 INTEGER NOT NULL,
    month                INTEGER NOT NULL,

    received_pcs         REAL NOT NULL,
    received_m3          REAL NOT NULL,
    
    file_name            TEXT NOT NULL,
    sheet_name           TEXT NOT NULL,
    sheet_name_orderable TEXT NOT NULL,
    row                  INTEGER,

    FOREIGN KEY (material_id) REFERENCES materials(id)
);
"""

def read_periods(workbooks: dict[str, Workbook]) -> Iterator[dict[str, int|dict[str, Worksheet]]]:
    for worksheet in workbooks['text'].worksheets:
        m = re.fullmatch(r"((0[1-9])|1[0-2])\.((19|20)[0-9]{2})", worksheet.title.strip(' .')) # parses dd.yyyy 01.1900-12.2099
        if not m:
            continue

        try:
            with log.prefix(f"parsing {worksheet.title}: "):
                yield {
                    "worksheet": {
                        "text": worksheet,
                        "formulas": workbooks["formulas"][worksheet.title]
                    },
                    "month": int(m.group(1)),
                    "year": int(m.group(3)),
                }
        except Exception as e:
            log.log(f'Exception raised when parsing: {worksheet.title}')
            raise e

def try_float(value, safe = False) -> float:
    if isinstance(value, float): return value
    if isinstance(value, int): return float(value)

    try:
        return float(value.replace(' ', '').replace(',', '.').replace(' ', ''))
    except Exception as e:
        if safe:
            return 0.0
        else:
            raise e

def normalize_text(value: Any) -> str|None:
    if value is None:
        return None
    s = str(value).strip()
    s = re.sub(r"\s+", " ", s) # squish
    return s

def handle_special_cases(row: Any, category: str = None, brand: str = None, worksheet: dict[str, Worksheet] = None):
    # a bunch of hardcoded special rules
    if category is None:
        record = row
        if record['IEPIRKUMA CENA']['text'] == 'Daniil':
            record['IEPIRKUMA CENA']['text'] = 0

        if record.get('BIEZUMS (m)') is None:
             record['BIEZUMS (m)'] = record['AUGSTUMS (m)']
        if record['AUGSTUMS (m)'] is None:
             record['AUGSTUMS (m)'] = record['BIEZUMS (m)']

        if record.get('ATLIKUMS (gb.)') is None:
            if (record.get('IEPRIEKŠĒJAIS ATLIKUMS (gb.)')):
                record['ATLIKUMS (gb.)'] = record['IEPRIEKŠĒJAIS ATLIKUMS (gb.)']
            else:
                record['ATLIKUMS (gb.)'] = {"text": None}
        if record.get('ATLIKUMS (m3)') is None:
            record['ATLIKUMS (m3)'] = {"text": None}
        if record.get('ATLIKUMS (m2)') is None:
            record['ATLIKUMS (m2)'] = {"text": None}

        return record
    else:
        if brand == 'BUVNIEKS':
            if category == 'Zāģmateriali':
                category = 'NEĒVELĒTI'
            if category == 'Ēvelēts' or category == 'Ēvelēti':
                category = 'ĒVELĒTI'
        if brand == 'KARTEX WOOD':
            brand = 'KARTEX'
        if brand == 'RONDI' and category == 'C24 no Zviedrijas':
            category = 'C24'
        return row, category, brand





def worksheet_records(worksheet: dict[str, Worksheet]) -> Iterator[defaultdict[str, str|int|dict[str, str|float|None]]]:
    for row in range(1, 10):
        headers = [normalize_text(cell.value or '') for cell in worksheet['text'][row]]
        if ('KVALITĀTE' in headers):
            break
    else:
        return

    brand = None
    category = None
    is_start_of_range = lambda cell: hasattr(cell, 'internal_value')
    while row < worksheet['text'].max_row:
        row += 1

        if 'KOPĀ VISS' in (worksheet['text'][row][1].value or ''): # table end
            break

        if (not brand or is_start_of_range(worksheet['text'][row][0])):
            brand = normalize_text(worksheet['text'][row][0].value or '') # new brand start

        category = category or worksheet['text'][row][1].value
        if 'KOPĀ' in (worksheet['text'][row][1].value or ''):
            category = None # end of category (possibly of the whole brand or the whole table. either way, skipping)
            continue

        if (not worksheet['text'][row][2].value): continue # empty line

        row, category, brand = handle_special_cases(row, category, brand, worksheet)

        record: defaultdict[str, str|dict] = defaultdict(lambda: None)
        record["brand"] = brand
        record["category"] = category
        record["row"] = row



        for col, header in enumerate(headers):
            text_cell: Cell = worksheet['text'][row][col]
            formula_cell: Cell = worksheet['formulas'][row][col]

            formula = formula_cell.value

            record[header] = {
                "text": normalize_text(text_cell.value),
                "formula": formula[1:] if formula is not None and isinstance(formula, str) and formula.startswith('=') and '+' in formula else None,

                "float": try_float(text_cell.value, True),
                "_text_cell": text_cell,
            }
        try:
            log.log(f'\tparsing {row}')
            yield handle_special_cases(record)
        except Exception as e:
            log.log(f'Exception raised when parsing: {worksheet['text'].title}:{text_cell.coordinate}')
            raise e

# same ID for repeated materials
def material_hash(material):
    data = f"{material['brand']}|{material['category']}|{material['quality']}|{material['height_m']}|{material['width_m']}|{material['length_m']}"
    return hashlib.sha256(data.encode("utf-8")).hexdigest()
def uuid():
    return str(uuid4())
def equals(left: float, right: float):
    return round(abs(round(left, 2) - round(right, 2)), 2) <= 0.01

def parse_sales(record):
    # pcs =1+2+3 (three separate records)
    # m3  =pcs*m3 (single record)
    # eur =1+2+3 (three separate records)
    pcs = record['PĀRDOTS (gb.)']
    eur = record['PĀRDOTS (€)']

    if pcs['formula'] is not None and eur['formula'] is not None:
        pcs = [try_float(v) for v in pcs['formula'].split('+')]
        eur = [try_float(v) for v in eur['formula'].split('+')]

        m3_per_pc = record['m3 (1 gb)']['float']
        m3 = [pc*m3_per_pc for pc in pcs]

        if (not equals(sum(m3), record['PĀRDOTS (m3)']['float'])):
            log.log('PĀRDOTS (m3) does not match formula')
        assert(len(pcs) == len(eur))

        return zip(pcs, m3, eur)

    if record['PĀRDOTS (gb.)']['float'] > 0:
        return ((record['PĀRDOTS (gb.)']['float'], record['PĀRDOTS (m3)']['float'], record['PĀRDOTS (€)']['float']),)
    else:
        return []
def parse_deliveries(record):
    # pcs =1+2+3 (three separate records)
    # eur =pcs*price (single record)
    pcs = record['SAŅEMAM (gb.)']

    if pcs['formula'] is not None:
        pcs = [try_float(v) for v in pcs['formula'].split('+')]

        m3_per_pc = record['m3 (1 gb)']['float']
        m3 = [pc*m3_per_pc for pc in pcs]

        assert(equals(sum(m3), record['SAŅEMAM (m3)']['float']))

        return zip(pcs, m3)

    if record['SAŅEMAM (gb.)']['float'] > 0:
        return ((record['SAŅEMAM (gb.)']['float'], record['SAŅEMAM (m3)']['float']),)
    else:
        return []




def upsert_many(connection: sqlite3.Connection, table: str, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        return

    columns = list(rows[0].keys())
    placeholders = ", ".join(["?"] * len(columns))
    column_list = ", ".join(columns)

    sql = f"INSERT INTO {table} ({column_list}) VALUES ({placeholders}) ON CONFLICT(id) DO NOTHING"
    connection.executemany(sql, ([row.get(c) for c in columns] for row in rows))



if os.path.exists('timber.sqlite'):
    os.remove('timber.sqlite')

sources = [p for p in Path("sources").glob("*.xlsx") if not p.name.startswith('~')]
with sqlite3.connect("timber.sqlite") as db:
    db.executescript(SCHEMA_SQL)

    for source in sources:
        workbooks: dict[str, Workbook] = {"text": load_workbook(source, data_only=True), "formulas": load_workbook(source, data_only=False)}

        for period in read_periods(workbooks):
            materials = dict()
            end_of_month_material_stats = list()
            sales = list()
            deliveries = list()
            for record in worksheet_records(period['worksheet']):
                with_source = lambda data: data | {
                    "file_name": source.name,
                    "sheet_name": period['worksheet']['text'].title,
                    "sheet_name_orderable": f"{period['year']}-{period['month']:02d}",
                    "row": record['row']
                }

                material = with_source({
                    "id": None,

                    "brand":     record['brand'],
                    "category":  record['category'],
                    "quality":   record['KVALITĀTE']['text'],

                    "height_m":  record['BIEZUMS (m)']['text'],
                    "width_m":   record['PLATUMS (m)']['text'],
                    "length_m":  record['GARUMS (m)']['text'],
                })
                material_id = material_hash(material)
                material['id'] = material_id
                materials[material_id] = material

                end_of_month_material_stats.append(with_source({
                    "id": uuid(),

                    "material_id":    material_id,
                    "year":           period['year'],
                    "month":          period['month'],

                    "price_purchase": record['IEPIRKUMA CENA']['text'],
                    "price_m3":       record['Cena m3']['text'],
                    "price_pcs":      record['Cena 1 gb']['text'],
                    "price_m2":       record['Cena m2']['text'],

                    "qty_pcs":        record['ATLIKUMS (gb.)']['text'],
                    "qty_m3":         record['ATLIKUMS (m3)']['text'],
                    "qty_m2":         record['ATLIKUMS (m2)']['text'],
                    "qty_eur":        record['ATLIKUMS (€)']['text'],
                }))


                for pcs, m3, eur in parse_sales(record):
                    sales.append(with_source({
                        "id": uuid(),

                        "material_id":   material_id,
                        "year":          period['year'],
                        "month":         period['month'],

                        "sale_pcs":      pcs,
                        "sale_m3":      m3,
                        "sale_eur":      eur,
                    }))
                for pcs, m3 in parse_deliveries(record):
                    deliveries.append(with_source({
                        "id": uuid(),

                        "material_id":   material_id,
                        "year":          period['year'],
                        "month":         period['month'],

                        "received_pcs":  pcs,
                        "received_m3":   m3,
                    }))



            # Upserts
            upsert_many(db, "materials", list(materials.values()))
            upsert_many(db, "end_of_month_material_stats", end_of_month_material_stats)
            upsert_many(db, "sales", sales)
            upsert_many(db, "deliveries", deliveries)

# TODO: ignore empty lines (see 07.2025)
# TODO: ignore records about written off sales